import numbers
from dataclasses import dataclass, field
from typing import Dict, List, Sequence, Set, Tuple

import pandas as pd
from openpyxl import load_workbook

# from .ambr import AmbrExcelParser
# import ambr
from .. import parsers


@dataclass(frozen=True)
class MeasurementParseRecord:

    loa_name: str
    mtype_name: str
    value_format: str
    data: List[List[numbers.Number]]
    x_unit_name: str
    y_unit_name: str
    # data source(s) within the file for this measurement...for tabular data, an iterable of
    # int row nums or string ranges, e.g. ('1-3', 5, 24). Used to construct helpful / precise
    # error messages
    src_ids: [Tuple[str, ...]]


@dataclass(frozen=True)
class ParseResult:

    series_data: Sequence[MeasurementParseRecord]
    # a human-readable identifier for the file portion(s) identified by each
    # MeasurementParseRecord's src_ids. For example, for tabular data, "row" is
    # often used
    record_src: str
    # True if the file contained at least one time value
    any_time: bool
    # True if a time value was found
    # corresponding to each MeasurementParseRecord
    # Since parsed values may be packed differently,
    # it's simpler for the parser to make this determination.
    has_all_times: bool

    # set of unique line or assay names found in the file
    line_or_assay_names: Set[str] = field(init=False)
    # True if every record parsed from file had associated units
    has_all_units: bool = field(init=False)
    # set of unique mtype identifiers (strings) found in the file
    mtypes: Set[str] = field(init=False)
    # set of unique unit names (strings) found in the file or implicit in the format
    units: Set[str] = field(init=False)

    def __post_init__(self):
        # compute unique line / assay names, units, measurement types from the parse records
        line_or_assay_names = set()
        units = set()
        mtypes = set()
        has_all_units = True
        for record in self.series_data:
            line_or_assay_names.add(record.loa_name)
            units.add(record.x_unit_name)
            units.add(record.y_unit_name)
            mtypes.add(record.mtype_name)
            if not record.x_unit_name or not record.y_unit_name:
                has_all_units = False
        # set member fields
        self.line_or_assay_names = frozenset(line_or_assay_names)
        self.has_all_units = has_all_units
        self.mtypes = frozenset(mtypes)
        self.units = frozenset(units)


class MultiSheetExcelParserMixin:
    def parse(self, file):
        wb = load_workbook(file, read_only=False, data_only=True)
        print("In parse(). workbook has %d sheets" % len(wb.worksheets))

        for worksheet in wb.worksheets:
            print(worksheet)

            for col in worksheet.iter_cols():
                print(col)
                for cell in col:
                    if cell.value is not None:
                        print(cell.value)
            # return self._parse_rows(worksheet.iter_rows())
        return None

    def parse_pd(self, file):

        wb = pd.read_excel(file, sheet_name=None)
        # print(wb)

        for name, sheet in wb.items():
            # print(name)
            # print(sheet)

            # for each measurement type
            for i in range(0, int(sheet.shape[1]), 2):
                # print(i)
                # print(sheet[sheet.columns[i:i+2]])

                two_cols = sheet[sheet.columns[i : i + 2]]

                if not two_cols.dropna().empty:
                    # print(two_cols)
                    # print(two_cols[two_cols.columns[1:2]])

                    # print('----------')
                    mapper = MeasurementMapper(name, two_cols)

                    # convert mapper.df into a openpyxl worksheet
                    # and then call parse_rows with worksheet.iterrows()

                    # call MeasurementParseRecord
                    # print(mapper.df)

                    for row in mapper.df.itertuples():
                        print(row.LineName)
                        # print(row.Index)

                        # create measurement records for one measurement type
                        mpr = MeasurementParseRecord(
                            loa_name=row.LineName,
                            mtype_name=row.MeasurementType,
                            data=row.Values,
                            y_unit_name=row.Units,
                            value_format="SCALAR",
                            x_unit_name="hours",  # assumed hours for time, included in EDD's bootstrap.json
                            src_ids=(row.Index + 1,),
                        )


class MeasurementMapper:

    loa_name: str
    mtype_name: str
    data: List[List[numbers.Number]]
    y_unit_name: str
    df: pd.DataFrame

    def __init__(self, sheet_name, df):

        loa_name = sheet_name
        mtype_name = df[df.columns[1:2]].columns.values[0]
        df["LineName"] = loa_name
        df.columns.values[0] = "Time"
        df.columns.values[1] = "Values"

        df["MeasurementType"] = mtype_name
        df["Units"] = "units"
        self.df = df


# mpm = MultiSheetExcelParserMixin()
# file_name = '/Users/somtirtharoy/Downloads/ABPDU_EDD_import/demo.xlsx'
# mpm.parse_pd(file_name)


file_name = "/Users/somtirtharoy/Downloads/ABPDU_EDD_import/demo.xlsx"

ap = parsers.AmbrExcelParser()
ap.parse(file_name)
